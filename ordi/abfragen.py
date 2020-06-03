

from datetime import date, datetime
import os

from flask import Blueprint, flash, g, redirect, render_template, request, url_for, send_file
from werkzeug.exceptions import abort
from sqlalchemy import func, distinct, or_, and_

from . import db
from ordi.auth import login_required
from ordi.models import *
from ordi.reqhelper import *
from ordi.values import *
from ordi.createpdf import *

bp = Blueprint('abfrage', __name__, url_prefix='/abfrage')


def dl_bericht(abfrage, tierhaltungen, kriterium1, kriterium2):
    html = render_template('abfragen/print.html', abfrage=abfrage, kriterium1=kriterium1, 
        kriterium2=kriterium2, tierhaltungen=tierhaltungen)
    filename = abfrage + "_bericht.pdf"
    path_and_filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads', filename)

    html2pdf_blank(html, path_and_filename)

    return path_and_filename


def dl_etiketten(abfrage, personen):
    html = render_template('abfragen/print_etiketten.html', personen=personen)
    filename = abfrage + "_etiketten.pdf"
    path_and_filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads', filename)

    html2pdf_blank(html, path_and_filename)

    return path_and_filename


def dl_excel(abfrage, tierhaltungen):
    filename = abfrage + ".xlsx"
    path_and_filename = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads', filename)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Behandlungsdatum"
    ws['B1'] = "Familienname"
    for tierhaltung in tierhaltungen:
        print(tierhaltung)
        ws.append([tierhaltung.Behandlung.datum.strftime("%d.%m.%Y"), tierhaltung.Person.familienname])
    wb.save(path_and_filename)

    return path_and_filename


def str_to_date(str_date):
    datum = None
    error = ""

    try:
        datum = datetime.strptime(str_date, '%d.%m.%Y').date()
    except:
        error += "Falsches Datumsformat: TT.MM.JJJJ erwartet."

    return datum, error


lst_abfragen = ["", "Arzneien", "Arzneimittel", "Behandlung", "Chipnummer", \
                "Diagnose", "EU-Passnummer", "Finanzamt", "Impfung", "Kontakte", \
                "Merkmal", "Postleitzahl", "Rasse", "Straße", "Tierart"]

@bp.route('/index', methods=('GET', 'POST'))
@login_required
def index():
    output = ""

    if(request.method == 'POST'):
        abfrage = request.form['abfrage']

        kriterium1 = request.form['kriterium1']
        try:
            kriterium2 = request.form['kriterium2']
        except:
            kriterium2 = ""

        twokriteria = 0

        if(request.form.get('kunde')):
            kunde = True
        else:
            kunde = False

        if(request.form.get('patient')):
            patient = True
        else:
            patient = False

        try:
            output = request.form['output']
        except:
            pass

        tierhaltungen = []

        if(len(abfrage) == 0 or len(kriterium1) == 0):
            error = ""
            if(len(abfrage) == 0):
                error += "Abfrage fehlt. "
            if(len(kriterium1) == 0):
                error += "Eingabe Textfeld1 fehlt. "
            flash(error)
            return render_template('abfragen/index.html', abfragen=lst_abfragen, 
                abfrage=abfrage, kriterium1=kriterium1, kriterium2=kriterium2, 
                twokriteria=twokriteria, kunde=kunde, patient=patient, 
                tierhaltungen=tierhaltungen, page_title="Abfragen")

        if(abfrage == "Behandlung" or abfrage == "Finanzamt" or abfrage == "Impfung"):
            error = ""
            twokriteria = 1

            if(len(kriterium2) == 0):
                error += "Eingabe Textfeld2 fehlt. "

            if(len(error) == 0):
                von_datum, error = str_to_date(kriterium1)

            if(len(error) == 0):
                bis_datum, error = str_to_date(kriterium2)

            if(len(error) > 0):
                flash(error)
                return render_template('abfragen/index.html', abfragen=lst_abfragen, 
                    abfrage=abfrage, kriterium1=kriterium1, kriterium2=kriterium2, 
                    twokriteria=twokriteria, kunde=kunde, patient=patient, 
                    tierhaltungen=tierhaltungen, page_title="Abfragen")

        if(abfrage == "Arzneien" or abfrage == "Arzneimittel" or abfrage == "Behandlung" or 
           abfrage == "Diagnose" or abfrage == "Finanzamt"):
            if(abfrage == "Finanzamt" and output == "excel"):
                tierhaltungen = db.session.query(Person, Behandlung) \
                    .join(Person.tierhaltungen) \
                    .join(Tierhaltung.tier) \
                    .join(Tier.behandlungen) \
                    .filter(Behandlung.datum >= von_datum, Behandlung.datum <= bis_datum, \
                            ~Behandlung.diagnose.contains('%Tel%'), \
                            ~Tier.merkmal.contains('%Abzeichen%')) \
                    .order_by(Behandlung.datum.asc()).all()

                path_and_filename = dl_excel(abfrage, tierhaltungen)
                return send_file(path_and_filename, as_attachment=True)
            else:
                query = db.session.query(Tierhaltung, Person, Tier) \
                    .join(Tierhaltung.person) \
                    .join(Tierhaltung.tier) \
                    .join(Tier.behandlungen) \
                    .order_by(Person.familienname.asc())
        elif(abfrage == "Impfung"):
            if(output == "etiketten"):
                personen = db.session.query(Person) \
                    .join(Person.tierhaltungen) \
                    .join(Tierhaltung.tier) \
                    .join(Tier.behandlungen) \
                    .join(Behandlung.impfungen) \
                    .filter(Person.kunde==kunde, Tier.patient==patient, \
                            Behandlung.datum >= von_datum, Behandlung.datum <= bis_datum) \
                    .order_by(Person.familienname.asc()).all()

                path_and_filename = dl_etiketten(abfrage, personen)
                return send_file(path_and_filename, as_attachment=True)
            else:
                query = db.session.query(Tierhaltung, Person, Tier) \
                    .join(Tierhaltung.person) \
                    .join(Tierhaltung.tier) \
                    .join(Tier.behandlungen) \
                    .join(Behandlung.impfungen) \
                    .order_by(Person.familienname.asc())
        else:
            query = db.session.query(Tierhaltung, Person, Tier) \
                .join(Tierhaltung.person) \
                .join(Tierhaltung.tier) \
                .order_by(Person.familienname.asc())

        if(abfrage == "Arzneien"):
            tierhaltungen = query.filter(Behandlung.arzneien.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Arzneimittel"):
            tierhaltungen = query.filter(Behandlung.arzneimittel.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Behandlung"):
            tierhaltungen = query.filter(Behandlung.datum >= von_datum, Behandlung.datum <= bis_datum, \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Chipnummer"):
            tierhaltungen = query.filter(Tier.chip_nummer.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Diagnose"):
            tierhaltungen = query.filter(Behandlung.diagnose.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "EU-Passnummer"):
            tierhaltungen = query.filter(Tier.eu_passnummer.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Finanzamt"):
            tierhaltungen = query.filter(Behandlung.datum >= von_datum, Behandlung.datum <= bis_datum, 
                                ~Behandlung.diagnose.contains('Tel%'), \
                                ~Tier.merkmal.contains('Abzeichen%')).all()
        elif(abfrage == "Impfung"):
            tierhaltungen = query.filter(Behandlung.datum >= von_datum, Behandlung.datum <= bis_datum, \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Kontakte"):
            tierhaltungen = query.filter(Person.kontakte.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Merkmal"):
            tierhaltungen = query.filter(Tier.merkmal.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Postleitzahl"):
            tierhaltungen = query.filter(Person.adr_plz.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Rasse"):
            tierhaltungen = query.filter(Tier.rasse.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Straße"):
            tierhaltungen = query.filter(Person.adr_strasse.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        elif(abfrage == "Tierart"):
            tierhaltungen = query.filter(Tier.tierart.like("%" + kriterium1 + "%"), \
                                         Person.kunde==kunde, Tier.patient==patient).all()
        else:
            tierhaltungen = []
    else:
        abfrage = ""
        kriterium1 = ""
        kriterium2 = ""
        twokriteria = 0
        kunde = 1
        patient = 1
        tierhaltungen = []

    if(output == "bericht-drucken"):
        path_and_filename = dl_bericht(abfrage, tierhaltungen, kriterium1, kriterium2)
        return send_file(path_and_filename, as_attachment=True)
    else:
        return render_template('abfragen/index.html', abfragen=lst_abfragen, 
            abfrage=abfrage, kriterium1=kriterium1, kriterium2=kriterium2, 
            twokriteria=twokriteria, kunde=kunde, patient=patient, 
            tierhaltungen=tierhaltungen, page_title="Abfragen")

